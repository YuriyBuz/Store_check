import streamlit as st
import pandas as pd
import requests
import random
import re
from io import BytesIO
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font

# URL вашої Google Таблиці
SHEET_URL = "https://docs.google.com/spreadsheets/d/16OIt-2jMpLGehrYGEcv9_QpEfzvuetsleB3vJRl7rBY/edit?gid=0#gid=0"

sys.path.insert(0, str(Path(__file__).parent))
from scrape_brands import (
    scrape_silpo, scrape_novus, scrape_varus, node_available,
    write_store_sheet, write_summary_sheet, check_data_quality,
)

st.set_page_config(
    page_title="Моніторинг Залишків UA",
    page_icon="🛒",
    layout="wide",
)

# ── ДОПОМІЖНІ ФУНКЦІЇ ─────────────────────────────────────

def make_excel(results, query, checked_at):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Зведення")
    write_summary_sheet(ws, results, query, checked_at)
    for store, products in results.items():
        ws = wb.create_sheet(store)
        if products:
            write_store_sheet(ws, products, store, query, checked_at)
        else:
            ws["A1"] = f'Товарів не знайдено за запитом "{query}" у мережі {store}.'
            ws["A1"].font = Font(bold=True, color="FF0000")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def to_float(val):
    try:
        return float(str(val).replace(" ", "").replace(",", "."))
    except (ValueError, AttributeError):
        return None


def to_df(products):
    rows = []
    for p in products:
        rows.append({
            "Товар":          p.get("name", ""),
            "SKU":            p.get("sku", ""),
            "Ціна (грн)":     to_float(p.get("price", "")),
            "Знижка":         "Так" if p.get("on_discount") else "Ні",
            "Зі знижкою":     to_float(p.get("discount_price")) if p.get("on_discount") else None,
            "Наявність":      "Є" if p.get("in_stock") is True else ("Немає" if p.get("in_stock") is False else "?"),
            "Мережа":         p.get("seller", ""),
            "Посилання":      p.get("url", ""),
        })
    return pd.DataFrame(rows)

@st.cache_data(ttl=600)
def load_config_from_gsheet(url):
    """Завантажує налаштування мереж та брендів з Google Таблиці. Кешується на 10 хвилин."""
    try:
        # Надійне формування посилання для експорту CSV, витягуємо ID документу
        doc_id_match = re.search(r'/d/([a-zA-Z0-9-_]+)', url)
        if doc_id_match:
            doc_id = doc_id_match.group(1)
            csv_url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv&gid=0"
        else:
            csv_url = url
            
        df = pd.read_csv(csv_url)
        
        # ВИПРАВЛЕННЯ КИРИЛИЦІ: 
        # Замінюємо українські літери Т, С, М на англійські T, C, M, щоб уникнути конфліктів розкладки
        df.columns = (
            df.columns.str.strip()
            .str.replace('Т', 'T')
            .str.replace('С', 'C')
            .str.replace('М', 'M')
        )
        
        # Перевірка чи завантажилась правильна таблиця
        if 'TC' not in df.columns:
            st.error(f"❌ Помилка: Не знайдено колонку 'TC'. Знайдені колонки: `{list(df.columns)}`")
            st.info("💡 **Підказка:** Можливо, ваша Google Таблиця закрита. Відкрийте таблицю -> Натисніть 'Поділитися' -> Змініть доступ на **'Усі, хто має посилання' (Anyone with the link)**.")
            return pd.DataFrame()

        df = df.fillna("") # Замінюємо порожні значення
        return df
    except Exception as e:
        st.error(f"Помилка завантаження таблиці: {e}")
        return pd.DataFrame()

# ── БІЧНА ПАНЕЛЬ (SIDEBAR) ────────────────────────────────

with st.sidebar:
    st.title("🛒 Моніторинг Залишків")
    st.caption("Автоматична перевірка наявності та цін товарів у супермаркетах")
    
    # Завантажуємо дані з Google Таблиці
    gsheet_df = load_config_from_gsheet(SHEET_URL)
    
    if gsheet_df.empty:
        # Зупиняємо виконання, помилка вже виведена у функції вище
        st.stop()
        
    # Витягуємо унікальні мережі та бренди з колонок, ігноруючи порожні рядки
    stores_from_sheet = [s.strip() for s in gsheet_df['TC'].unique() if str(s).strip()]
    brands_from_sheet = [b.strip() for b in gsheet_df['TM'].unique() if str(b).strip()]
    
    # Створюємо словник Бренд -> Категорія для формування точного пошуку
    brand_category_map = {}
    for _, row in gsheet_df.iterrows():
        b = str(row.get('TM', '')).strip()
        # Використовуємо strip для назви колонки категорії
        c = str(row.get('котегорія продукту', '')).strip() 
        if b and b not in brand_category_map:
            brand_category_map[b] = c

    st.divider()
    st.subheader("1. Оберіть мережі")
    st.caption("Мережі завантажено з Google Таблиці")
    
    selected_stores = []
    for store in stores_from_sheet:
        if st.checkbox(store, value=True):
            selected_stores.append(store)

    st.divider()
    st.subheader("2. Бренд та Категорія")
    st.caption("Бренди завантажено з Google Таблиці")

    selected_brand = st.selectbox(
        label="Бренд (ТМ):",
        options=brands_from_sheet,
    )

    # Автоматично підтягуємо категорію з таблиці для обраного бренду
    auto_category = brand_category_map.get(selected_brand, "")
    if auto_category:
        st.info(f"Категорія з таблиці: **{auto_category}**")
        
    # Формуємо фінальний пошуковий запит
    # ВАЖЛИВО: Сайти погано реагують на складні запити типу "Соус Bonsai". 
    # Тому ми шукаємо лише за назвою бренду (як реальні користувачі).
    search_query = selected_brand

    st.divider()
    go = st.button("🔎  Знайти товари", type="primary", use_container_width=True)

# ── ГОЛОВНИЙ ЕКРАН ────────────────────────────────────────

st.title("Моніторинг Залишків та Цін")

if not go:
    st.info(
        "👈 Оберіть мережі та бренд у бічній панелі (всі налаштування підтягнуто з вашої онлайн-таблиці), "
        "а потім натисніть **Знайти товари**.\n\n"
        "Інструмент просканує сайти обраних супермаркетів і покаже актуальні ціни, знижки та наявність товарів."
    )
    st.stop()

if not selected_brand.strip():
    st.error("Будь ласка, оберіть назву бренду.")
    st.stop()

if not selected_stores:
    st.error("Будь ласка, оберіть хоча б одну мережу для пошуку.")
    st.stop()

# ── ПРОЦЕС ПАРСИНГУ ───────────────────────────────────────

has_node   = node_available()
session    = requests.Session()
results    = {}
metas      = {}
checked_at = datetime.now().strftime("%d.%m.%Y %H:%M")

st.write(f"### 🔍 Пошук за запитом: `{search_query}`")

for store in selected_stores:
    with st.status(f"Сканування {store}...", expanded=True) as s:
        meta = {}
        store_lower = store.lower()
        
        # Гнучка перевірка імені мережі, щоб уникнути проблем типу "Сильпо" vs "Сільпо"
        if "сільпо" in store_lower or "сильпо" in store_lower:
            products = scrape_silpo(search_query, session, has_node, log_fn=st.write, meta=meta)
        elif "novus" in store_lower:
            products = scrape_novus(search_query, session, log_fn=st.write, meta=meta)
        elif "varus" in store_lower:
            products = scrape_varus(search_query, session, log_fn=st.write, meta=meta)
        else:
            st.warning(f"Мережа '{store}' є в таблиці, але для неї ще не підключено алгоритм сканування.")
            products = []

        results[store] = products
        metas[store]   = meta
        
        n      = len(products)
        in_n   = sum(1 for p in products if p["in_stock"] is True)
        disc_n = sum(1 for p in products if p.get("on_discount"))

        if n > 0:
            s.update(
                label=f"✅ {store}: знайдено {n} товарів — {in_n} в наявності, {disc_n} зі знижкою",
                state="complete",
                expanded=False,
            )
        else:
            s.update(
                label=f"⚠️ {store}: товарів не знайдено за запитом '{search_query}'",
                state="error",
                expanded=True,
            )

# ── ПЕРЕВІРКА РЕЗУЛЬТАТІВ ─────────────────────────────────

total = sum(len(v) for v in results.values())
if total == 0:
    st.warning(f"Товарів за запитом **{search_query}** не знайдено в жодній з обраних мереж.")
    st.stop()

# ── МЕТРИКИ ───────────────────────────────────────────────

st.divider()
total_in   = sum(sum(1 for p in v if p["in_stock"] is True)  for v in results.values())
total_out  = sum(sum(1 for p in v if p["in_stock"] is False) for v in results.values())
total_disc = sum(sum(1 for p in v if p.get("on_discount"))   for v in results.values())

c1, c2, c3, c4 = st.columns(4)
c1.metric("Всього товарів", total)
c2.metric("В наявності",    total_in)
c3.metric("Відсутні",       total_out)
c4.metric("Зі знижкою",     total_disc)

# ── ПЕРЕВІРКА ЯКОСТІ ДАНИХ ────────────────────────────────

st.divider()
for store, products in results.items():
    if not products: continue
    
    warnings = check_data_quality(products)
    data_ok  = len(warnings) == 0

    if not data_ok:
        with st.expander(f"⚠️ Увага: перевірка даних — {store}", expanded=True):
            st.warning(f"Виявлено {len(warnings)} підозрілих товарів (наприклад, ціна 0 грн):")
            for msg, url in warnings:
                if url: st.caption(f"• {msg} — [Відкрити на сайті ↗]({url})")
                else: st.caption(f"• {msg}")

# ── ТАБЛИЦІ (ВКЛАДКИ) ─────────────────────────────────────

st.divider()

stores_with_data = [s for s in results if results[s]]
tab_labels = stores_with_data + ["📋 Зведення"]
tabs = st.tabs(tab_labels)

for i, store in enumerate(stores_with_data):
    with tabs[i]:
        df = to_df(results[store])

        # Фільтри для таблиці
        col_search, col_stock, col_disc = st.columns([3, 1, 1])
        with col_search:
            search_text = st.text_input(
                "Пошук за назвою",
                key=f"search_{store}",
                placeholder="Введіть текст для фільтрації...",
                label_visibility="collapsed",
            )
        with col_stock:
            stock_filter = st.selectbox(
                "Наявність",
                ["Всі", "Тільки Є", "Немає"],
                key=f"stock_{store}",
            )
        with col_disc:
            disc_filter = st.selectbox(
                "Знижка",
                ["Всі", "Так", "Ні"],
                key=f"disc_{store}",
            )

        filtered = df.copy()
        if search_text:
            filtered = filtered[filtered["Товар"].str.contains(search_text, case=False, na=False)]
        if stock_filter == "Тільки Є":
            filtered = filtered[filtered["Наявність"] == "Є"]
        elif stock_filter == "Немає":
            filtered = filtered[filtered["Наявність"] == "Немає"]
        if disc_filter != "Всі":
            filtered = filtered[filtered["Знижка"] == disc_filter]

        st.caption(f"Показано {len(filtered)} з {len(df)} товарів у мережі {store}")

        st.dataframe(
            filtered,
            use_container_width=True,
            hide_index=True,
            height=500,
            column_config={
                "Ціна (грн)": st.column_config.NumberColumn("Звичайна ціна", format="%.2f"),
                "Зі знижкою": st.column_config.NumberColumn("Акційна ціна", format="%.2f"),
                "Посилання": st.column_config.LinkColumn("Лінк", display_text="Відкрити ↗"),
            },
        )

with tabs[-1]:
    rows = []
    for store, prods in results.items():
        if not prods: continue
        in_n   = sum(1 for p in prods if p["in_stock"] is True)
        out_n  = sum(1 for p in prods if p["in_stock"] is False)
        disc_n = sum(1 for p in prods if p.get("on_discount"))
        rows.append({
            "Мережа":        store,
            "Всього":        len(prods),
            "В наявності":   in_n,
            "Відсутні":      out_n,
            "Зі знижкою":    disc_n,
        })
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ── ЕКСПОРТ В EXCEL ───────────────────────────────────────

st.divider()
ts    = datetime.now().strftime("%Y%m%d_%H%M")
fname = f"Залишки_{search_query.replace(' ', '_')}_{ts}.xlsx"

st.download_button(
    label="📥  Завантажити звіт в Excel",
    data=make_excel(results, search_query, checked_at),
    file_name=fname,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)
