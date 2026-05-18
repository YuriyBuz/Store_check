#!/usr/bin/env python3
"""
Універсальний парсер (Епіцентр, Eva, Organic Market, Сільпо, Novus, Varus)
Містить ОРИГІНАЛЬНІ алгоритми з вашого файлу + нові.
"""

import math
import requests
import subprocess
import time
import json
import re
import sys
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from bs4 import BeautifulSoup

try:
    import cloudscraper
except ImportError:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

# ─────────────────────────────────────────────────────────
#  НАЛАШТУВАННЯ ТА СТИЛІ
# ─────────────────────────────────────────────────────────
MAX_PAGES = 10

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "uk-UA,uk;q=0.9,ru;q=0.8,en;q=0.6",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
}

XLSX_COLS = [
    ("#",               5),
    ("Товар",          55),
    ("Артикул / SKU",  18),
    ("Ціна (грн)",     14),
    ("Знижка",         12),
    ("Ціна зі знижкою", 18),
    ("Наявність",      12),
    ("Мережа",         20),
    ("Посилання",      65),
]

C_HEADER    = "10B981" 
C_IN_STOCK  = "D1FAE5"
C_OUT_STOCK = "FEE2E2"
C_UNKNOWN   = "F1F5F9"
C_TITLE     = "047857"
C_DISCOUNT  = "FFEDD5"

# ─────────────────────────────────────────────────────────
#  ОРИГІНАЛЬНІ БАЗОВІ ФУНКЦІЇ З ВАШОГО КОДУ
# ─────────────────────────────────────────────────────────

def fetch(url, session):
    try:
        if 'cloudscraper' in sys.modules:
            scraper = cloudscraper.create_scraper()
            r = scraper.get(url, headers=REQUEST_HEADERS, timeout=25, allow_redirects=True)
        else:
             r = session.get(url, headers=REQUEST_HEADERS, timeout=25, allow_redirects=True)
        r.raise_for_status()
        r.encoding = 'utf-8'
        return BeautifulSoup(r.text, "html.parser"), r.text
    except requests.RequestException as e:
        return None, str(e)

def _node_cmd():
    for cmd in ("node", "nodejs"):
        try:
            if subprocess.run([cmd, "--version"], capture_output=True, timeout=5).returncode == 0:
                return cmd
        except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
            pass
    return None

def node_available():
    return _node_cmd() is not None

def decode_nuxt(script_text, log_fn=print):
    import tempfile, os
    cmd = _node_cmd()
    if not cmd:
        return None
    js = "const window={};\n" + script_text + "\nprocess.stdout.write(JSON.stringify(window.__NUXT__||null));"
    tmp = None
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".js", delete=False) as f:
            f.write(js)
            tmp = f.name
        result = subprocess.run([cmd, tmp], capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout:
            return json.loads(result.stdout)
    except Exception:
        pass
    finally:
        if tmp:
            try: os.unlink(tmp)
            except OSError: pass
    return None

def _resolve_nuxt(arr, idx, depth=0):
    if depth > 30 or not isinstance(idx, int) or idx < 0 or idx >= len(arr):
        return None
    val = arr[idx]
    if isinstance(val, list) and len(val) == 2 and isinstance(val[0], str) and isinstance(val[1], int):
        return _resolve_nuxt(arr, val[1], depth + 1)
    if isinstance(val, dict):
        return {k: (_resolve_nuxt(arr, v, depth + 1) if isinstance(v, int) else v) for k, v in val.items()}
    if isinstance(val, list):
        return [_resolve_nuxt(arr, v, depth + 1) if isinstance(v, int) else v for v in val]
    return val

# ─────────────────────────────────────────────────────────
#  1. ОРИГІНАЛЬНИЙ ПАРСЕР ЕПІЦЕНТР (З ВАШОГО АРХІВУ)
# ─────────────────────────────────────────────────────────
def find_epicenter_brand_url(brand, session, log_fn=print):
    brand_slug = brand.lower().replace(" ", "-")
    candidates = [
        f"https://epicentrk.ua/brands/{brand_slug}.html",
        f"https://epicentrk.ua/ua/brands/{brand_slug}.html",
        f"https://epicentrk.ua/brands/{urllib.parse.quote(brand.lower())}.html",
        f"https://epicentrk.ua/ua/brands/{urllib.parse.quote(brand.lower())}.html",
    ]
    for url in candidates:
        try:
            r = session.get(url, headers=REQUEST_HEADERS, timeout=15, allow_redirects=True)
            if r.status_code == 200 and "window.__NUXT__" in r.text:
                return r.url.split("?")[0]
        except Exception:
            pass
    return None

def _epicenter_page_url(brand_base, page_num, use_brand_page):
    if page_num == 1: return brand_base
    return f"{brand_base}?PAGEN_1={page_num}" if use_brand_page else f"{brand_base}&page={page_num}"

def _parse_epicenter_products(raw_products):
    page_products = []
    for p in raw_products:
        name      = p.get("name_ua") or p.get("name_ru") or ""
        sku       = str(p.get("id") or "")
        url_p     = p.get("url") or ""
        price_now = p.get("price") or 0
        price_old = p.get("price_old") or 0
        on_discount    = bool(price_old and price_old > price_now)
        regular_price  = str(price_old if on_discount else price_now)
        discount_price = str(price_now) if on_discount else ""
        avail = p.get("avail")
        if avail == 100: in_stock = True
        elif avail == 200: in_stock = "expected"
        elif avail in (350, 400): in_stock = False
        else: in_stock = None
        seller = p.get("seller") or "Epicenter"
        if name:
            page_products.append({
                "name": name, "sku": sku,
                "price": regular_price, "on_discount": on_discount,
                "discount_price": discount_price,
                "url": url_p, "in_stock": in_stock, "seller": seller,
            })
    return page_products

def _fetch_epicenter_page(page_num, brand_base, use_brand_page, session):
    url = _epicenter_page_url(brand_base, page_num, use_brand_page)
    soup, raw = fetch(url, session)
    if not soup: return page_num, None
    for script in soup.find_all("script"):
        txt = script.string or ""
        if "window.__NUXT__" in txt:
            nuxt_data = decode_nuxt(txt)
            if nuxt_data:
                try:
                    raw_prods = nuxt_data["state"]["products"]["products"]
                    return page_num, _parse_epicenter_products(raw_prods)
                except (KeyError, TypeError): pass
    return page_num, None

def scrape_epicenter(brand, session, has_node, log_fn=print, meta=None):
    if not has_node:
        log_fn("  ЕПІЦЕНТР: Node.js не встановлено. Парсинг неможливий.")
        return []
    log_fn(f"ЕПІЦЕНТР: шукаємо '{brand}'...")
    brand_base = find_epicenter_brand_url(brand, session, log_fn)
    if not brand_base:
        brand_enc = requests.utils.quote(brand)
        brand_base = f"https://epicentrk.ua/ua/search/?q={brand_enc}&per-page=60"
    use_brand_page = "brands" in brand_base
    soup, raw = fetch(_epicenter_page_url(brand_base, 1, use_brand_page), session)
    if not soup: return []
    nuxt_data = None
    for script in soup.find_all("script"):
        txt = script.string or ""
        if "window.__NUXT__" in txt:
            nuxt_data = decode_nuxt(txt, log_fn=log_fn)
            break
    if not nuxt_data: return []
    try:
        pagination  = nuxt_data["data"][0]["params"]["pagination"]
        total_pages = pagination.get("pages", 1)
    except (KeyError, IndexError, TypeError):
        total_pages = 1
    try:
        page1_products = _parse_epicenter_products(nuxt_data["state"]["products"]["products"])
    except (KeyError, TypeError):
        page1_products = []
    all_products = list(page1_products)
    if total_pages > 1:
        remaining = list(range(2, min(total_pages, MAX_PAGES) + 1))
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {executor.submit(_fetch_epicenter_page, p, brand_base, use_brand_page, session): p for p in remaining}
            page_results = {}
            for future in as_completed(futures):
                page_num, products = future.result()
                page_results[page_num] = products or []
        for p in remaining:
            all_products.extend(page_results.get(p, []))
    log_fn(f"  ЕПІЦЕНТР: знайдено {len(all_products)} товарів.")
    return all_products


# ─────────────────────────────────────────────────────────
#  2. ОРИГІНАЛЬНИЙ ПАРСЕР EVA (З ВАШОГО АРХІВУ)
# ─────────────────────────────────────────────────────────
def find_eva_brand_id(brand_name, session):
    try:
        r = session.get("https://api.eva.ua/v1/ua/api/brands", headers={**REQUEST_HEADERS, "Accept": "application/json"}, timeout=15)
        data = r.json()
        for group_brands in data.get("data", {}).get("groups", {}).values():
            for b in group_brands:
                if b.get("title", "").lower() == brand_name.lower():
                    m = re.search(r"brnd-(\d+)", b.get("url", ""))
                    if m: return m.group(1), b["title"]
    except Exception: pass
    return None, None

def parse_eva_nuxt_payload(html_text, brand_id):
    soup = BeautifulSoup(html_text, "html.parser")
    tag = soup.find("script", {"type": "application/json"})
    if not tag: return None
    try: arr = json.loads(tag.string or "")
    except Exception: return None
    brand_key = f"brnd-brnd-{brand_id}"
    for el in arr:
        if isinstance(el, dict) and brand_key in el:
            idx = el[brand_key]
            if isinstance(idx, int) and idx > 0:
                return _resolve_nuxt(arr, idx)
    return None

def _parse_eva_products(brand_data):
    page_products = []
    for p in brand_data.get("hits", []):
        name   = p.get("name") or ""
        sku    = str(p.get("sku") or "")
        price  = p.get("price") or 0
        final  = p.get("final_price") or p.get("special_price") or price
        on_discount    = bool(final and price and final < price)
        regular_price  = str(price)
        discount_price = str(final) if on_discount else ""
        stock       = p.get("stock") or {}
        in_stock    = stock.get("is_in_stock")
        url_p       = f"https://eva.ua/ua/search/?q={sku}" if sku else ""
        seller      = "EVA"
        if name:
            page_products.append({
                "name": name, "sku": sku,
                "price": regular_price, "on_discount": on_discount,
                "discount_price": discount_price,
                "url": url_p, "in_stock": in_stock, "seller": seller,
            })
    return page_products

def _fetch_eva_page(page_num, base_url, brand_id, session):
    url = f"{base_url}?p={page_num}" if page_num > 1 else base_url
    _, raw = fetch(url, session)
    if not raw: return page_num, None
    brand_data = parse_eva_nuxt_payload(raw, brand_id)
    if not brand_data: return page_num, None
    return page_num, _parse_eva_products(brand_data)

def scrape_eva(brand, session, log_fn=print, meta=None):
    log_fn(f"EVA: шукаємо '{brand}'...")
    brand_id, found_title = find_eva_brand_id(brand, session)
    if not brand_id:
        log_fn(f"  EVA: Бренд '{brand}' не знайдено.")
        return []
    base_url = f"https://eva.ua/ua/brnd-{brand_id}/"
    _, raw = fetch(base_url, session)
    if not raw: return []
    brand_data = parse_eva_nuxt_payload(raw, brand_id)
    if not brand_data: return []
    site_total = brand_data.get("total")
    page1_products = _parse_eva_products(brand_data)
    per_page    = len(page1_products) if page1_products else 40
    total_pages = math.ceil(int(site_total) / per_page) if site_total and per_page else 1
    all_products = list(page1_products)
    if total_pages > 1:
        remaining = list(range(2, min(total_pages, MAX_PAGES) + 1))
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {executor.submit(_fetch_eva_page, p, base_url, brand_id, session): p for p in remaining}
            page_results = {}
            for future in as_completed(futures):
                page_num, products = future.result()
                page_results[page_num] = products or []
        for p in remaining:
            all_products.extend(page_results.get(p, []))
    log_fn(f"  EVA: знайдено {len(all_products)} товарів.")
    return all_products


# ─────────────────────────────────────────────────────────
#  3. ОРИГІНАЛЬНИЙ ПАРСЕР ORGANIC MARKET (З ВАШОГО АРХІВУ)
# ─────────────────────────────────────────────────────────
def _organic_solve_challenge(session, url):
    r = session.get(url, headers=REQUEST_HEADERS, timeout=20, allow_redirects=True)
    if len(r.text) < 2000 and "challenge_passed" in r.text:
        m = re.search(r'defaultHash\s*=\s*"([a-f0-9]+)"', r.text)
        if m:
            session.cookies.set("challenge_passed", m.group(1), domain="organic-market.com.ua")
            r = session.get(url, headers=REQUEST_HEADERS, timeout=20, allow_redirects=True)
    return r

def find_organic_brand_url(brand, session, log_fn=print):
    slug = brand.lower().replace(" ", "-")
    candidates = [f"https://organic-market.com.ua/ru/{slug}/", f"https://organic-market.com.ua/ua/{slug}/"]
    for url in candidates:
        try:
            r = _organic_solve_challenge(session, url)
            if r.status_code == 200 and "catalogCard-box" in r.text:
                return r.url
        except Exception: pass
    return None

def _parse_organic_page(html):
    soup = BeautifulSoup(html, "html.parser")
    products = []
    for card in soup.find_all("div", class_="catalogCard-box"):
        title_a = card.find(class_="catalogCard-title")
        if not title_a: continue
        link = title_a.find("a")
        if not link: continue
        name    = link.get_text(strip=True)
        url_p   = "https://organic-market.com.ua" + link["href"] if link.get("href", "").startswith("/") else link.get("href", "")
        sku     = str(card.get("data-id", ""))
        old_el  = card.find(class_="catalogCard-oldPrice")
        new_el  = card.find(class_="catalogCard-price")
        def parse_price(txt):
            digits = re.sub(r"[^\d.]", "", txt.replace(",", ".").replace(" ", ""))
            try: return float(digits)
            except ValueError: return 0.0
        old_price = parse_price(old_el.get_text(strip=True) if old_el else "")
        new_price = parse_price(new_el.get_text(strip=True) if new_el else "")
        if old_price and old_price > new_price:
            on_discount, regular_price, discount_price = True, str(old_price), str(new_price)
        else:
            on_discount, regular_price, discount_price = False, str(new_price or old_price), ""
        in_stock = bool(card.find(class_=re.compile(r"j-buy-button-add")))
        products.append({
            "name": name, "sku": sku, "price": regular_price, "on_discount": on_discount,
            "discount_price": discount_price, "url": url_p, "in_stock": in_stock, "seller": "Organic Market",
        })
    return products

def _organic_total_pages(html):
    soup = BeautifulSoup(html, "html.parser")
    pager = soup.find("nav", class_="pager")
    if not pager: return 1
    page_nums = []
    for a in pager.find_all(class_="pager__item"):
        try: page_nums.append(int(a.get_text(strip=True)))
        except ValueError: pass
    return max(page_nums) if page_nums else 1

def _fetch_organic_page(page_num, brand_base, session):
    url = brand_base if page_num == 1 else f"{brand_base.rstrip('/')}/filter/page={page_num}/"
    try:
        r = _organic_solve_challenge(session, url)
        if r.status_code == 200: return page_num, _parse_organic_page(r.text)
    except Exception: pass
    return page_num, None

def scrape_organic(brand, session, log_fn=print, meta=None):
    log_fn(f"Organic Market: шукаємо '{brand}'...")
    brand_base = find_organic_brand_url(brand, session, log_fn)
    if not brand_base: return []
    try:
        r = _organic_solve_challenge(session, brand_base)
        if r.status_code != 200: return []
    except Exception: return []
    total_pages = _organic_total_pages(r.text)
    all_products = list(_parse_organic_page(r.text))
    if total_pages > 1:
        remaining = list(range(2, min(total_pages, MAX_PAGES) + 1))
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {executor.submit(_fetch_organic_page, p, brand_base, session): p for p in remaining}
            page_results = {}
            for future in as_completed(futures):
                page_num, products = future.result()
                page_results[page_num] = products or []
        for p in remaining:
            all_products.extend(page_results.get(p, []))
    log_fn(f"  Organic: знайдено {len(all_products)} товарів.")
    return all_products


# ─────────────────────────────────────────────────────────
#  4. ДОДАТКОВІ ПАРСЕРИ (СІЛЬПО ТА ZAKAZ) - Схильні до блокування Cloudflare
# ─────────────────────────────────────────────────────────
def extract_products_from_json(node, found_products):
    if isinstance(node, dict):
        if "price" in node and ("title" in node or "name" in node or "name_ua" in node):
            if "sku" in node or "id" in node or "ean" in node or "slug" in node:
                sku = str(node.get("id") or node.get("sku") or node.get("ean") or node.get("slug") or "")
                if sku and sku not in found_products: found_products[sku] = node
        for v in node.values(): extract_products_from_json(v, found_products)
    elif isinstance(node, list):
        for v in node: extract_products_from_json(v, found_products)

def scrape_silpo(query, session, has_node, log_fn=print, meta=None):
    log_fn(f"Сільпо: шукаємо '{query}'...")
    query_enc = requests.utils.quote(query)
    url = f"https://silpo.ua/search?find={query_enc}"
    soup, raw_text = fetch(url, session)
    if not soup:
        log_fn(f"  Сільпо: Помилка доступу. Блокування Cloudflare (403).")
        return []
    
    products_data = {}
    tag = soup.find("script", id="__NUXT_DATA__")
    if tag:
        try:
            arr = json.loads(tag.string)
            resolved = [_resolve_nuxt(arr, i) for i in range(min(15, len(arr)))]
            extract_products_from_json(resolved, products_data)
        except Exception: pass
    
    final_products = []
    for p in products_data.values():
        name = p.get("title") or p.get("name") or p.get("name_ua") or ""
        sku = str(p.get("id") or p.get("sku") or p.get("slug") or p.get("ean") or "")
        price = float(p.get("price", 0))
        old_price = float(p.get("oldPrice") or p.get("old_price") or p.get("price_old") or 0)
        on_discount = old_price > price
        reg_price = old_price if on_discount else price
        disc_price = price if on_discount else 0
        in_stock = str(p.get("status")) != "out_of_stock"
        url_p = f"https://silpo.ua/product/{p.get('slug')}" if p.get("slug") else url
        final_products.append({
            "name": name, "sku": sku, "price": str(reg_price), "on_discount": on_discount,
            "discount_price": str(disc_price) if on_discount else "", "url": url_p, "in_stock": in_stock, "seller": "Сільпо"
        })
    log_fn(f"  Сільпо: знайдено {len(final_products)} товарів.")
    return final_products

def scrape_zakaz(retailer, query, session, log_fn=print, meta=None):
    log_fn(f"{retailer.title()}: шукаємо '{query}'...")
    query_enc = requests.utils.quote(query)
    base_url = f"https://{retailer}.zakaz.ua/uk/search/?q={query_enc}"
    soup, raw_text = fetch(base_url, session)
    if not soup:
        log_fn(f"  {retailer.title()}: Помилка доступу. Блокування Cloudflare (403).")
        return []
    products_data = {}
    script = soup.find("script", id="__NEXT_DATA__")
    if script:
        try:
            data = json.loads(script.string)
            extract_products_from_json(data, products_data)
        except Exception: pass
    final_products = []
    for p in products_data.values():
        name = p.get("title") or p.get("name") or ""
        sku = str(p.get("ean") or p.get("id") or "")
        price_raw = float(p.get("price", 0))
        if price_raw > 1000: price_raw = price_raw / 100.0
        old_price_raw = float(p.get("old_price") or p.get("discount_price") or 0)
        if old_price_raw > 1000: old_price_raw = old_price_raw / 100.0
        on_discount = bool(old_price_raw and old_price_raw > price_raw)
        reg_price = old_price_raw if on_discount else price_raw
        disc_price = price_raw if on_discount else 0
        in_stock = p.get("in_stock", True)
        url_p = f"https://{retailer}.zakaz.ua/uk/products/{sku}/" if sku else base_url
        final_products.append({
            "name": name, "sku": sku, "price": str(reg_price), "on_discount": on_discount,
            "discount_price": str(disc_price) if on_discount else "", "url": url_p, "in_stock": in_stock, "seller": retailer.title()
        })
    log_fn(f"  {retailer.title()}: знайдено {len(final_products)} товарів.")
    return final_products

def scrape_novus(query, session, log_fn=print, meta=None):
    return scrape_zakaz("novus", query, session, log_fn, meta)

def scrape_varus(query, session, log_fn=print, meta=None):
    return scrape_zakaz("varus", query, session, log_fn, meta)

# ─────────────────────────────────────────────────────────
#  ЕКСПОРТ EXCEL
# ─────────────────────────────────────────────────────────
def check_data_quality(products):
    warnings = []
    for p in products:
        name = p.get("name", "")[:50]
        url  = p.get("url", "")
        try: price = float(str(p.get("price", 0)).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError): price = 0
        try: disc = float(str(p.get("discount_price", 0) or 0).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError): disc = 0

        if price == 0 and p.get("in_stock") is not False:
            warnings.append((f"Ціна 0 грн: {name}", url))
        if p.get("on_discount") and disc >= price:
            warnings.append((f"Знижка >= звичайної ціни: {name}", url))
    return warnings

def _fmt_price(raw):
    digits = re.sub(r"[^\d.]", "", str(raw).replace(",", "."))
    try: return "{:,.2f}".format(float(digits)).replace(",", " ")
    except ValueError: return str(raw)

def write_store_sheet(ws, products, store_name, query, checked_at):
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor=C_HEADER)
    in_fill  = PatternFill("solid", fgColor=C_IN_STOCK)
    out_fill = PatternFill("solid", fgColor=C_OUT_STOCK)
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    last_col = get_column_letter(len(XLSX_COLS))
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value     = f"{store_name}  ·  Пошук: {query}  ·  Перевірено: {checked_at}"
    t.font      = Font(bold=True, size=13, color=C_TITLE)
    t.alignment = center
    ws.row_dimensions[1].height = 26

    for ci, (col_name, col_width) in enumerate(XLSX_COLS, 1):
        c = ws.cell(row=2, column=ci, value=col_name)
        c.font, c.fill, c.alignment = hdr_font, hdr_fill, center
        ws.column_dimensions[get_column_letter(ci)].width = col_width
    ws.freeze_panes = "A3"

    disc_fill = PatternFill("solid", fgColor=C_DISCOUNT)
    aligns = [center, left, center, center, center, center, center, center, left]
    for i, p in enumerate(products, 1):
        row = i + 2
        status, stock_fill = ("Є", in_fill) if p["in_stock"] else ("Немає", out_fill)
        on_disc = p.get("on_discount", False)
        vals = [
            i, p.get("name", ""), p.get("sku", ""),
            _fmt_price(p.get("price", "")), "Так" if on_disc else "Ні",
            _fmt_price(p.get("discount_price", "")) if on_disc else "",
            status, p.get("seller", ""), p.get("url", "")
        ]
        for ci, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.alignment = aln
            if ci in (5, 6) and on_disc: c.fill = disc_fill
            elif ci == 7: c.fill = stock_fill

def write_summary_sheet(ws, results, query, checked_at):
    hdr_font, hdr_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor=C_HEADER)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"Зведення  ·  Пошук: {query}  ·  {checked_at}"
    t.font, t.alignment = Font(bold=True, size=13, color=C_TITLE), center
    ws.row_dimensions[1].height = 26

    headers = ["Мережа", "Всього", "В наявності", "Відсутні", "Зі знижкою", "Невідомо"]
    widths  = [20, 14, 14, 14, 14, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font, c.fill, c.alignment = hdr_font, hdr_fill, center
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (store, products) in enumerate(results.items(), 3):
        in_n   = sum(1 for p in products if p["in_stock"] is True)
        out_n  = sum(1 for p in products if p["in_stock"] is False)
        disc_n = sum(1 for p in products if p.get("on_discount"))
        unk_n  = sum(1 for p in products if p["in_stock"] is None)
        
        for ci, val in enumerate([store, len(products), in_n, out_n, disc_n, unk_n], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = center
            if ci == 3 and in_n > 0: c.fill = PatternFill("solid", fgColor=C_IN_STOCK)
            if ci == 4 and out_n > 0: c.fill = PatternFill("solid", fgColor=C_OUT_STOCK)
            if ci == 5 and disc_n > 0: c.fill = PatternFill("solid", fgColor=C_DISCOUNT)
